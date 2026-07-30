#!/usr/bin/env python3
"""Phase 1 Audit: Comprehensive product catalog audit → CURRENT_PRODUCTS_AUDIT.xlsx"""

import re
import os
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system("pip3 install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path("/Users/aniket/Desktop/geetai-template-engine")
PRODUCTS_DIR = BASE / "src" / "data" / "products"
IMAGES_DIR = BASE / "public" / "images"
OUTPUT_DIR = BASE / "output"
SCRIPT_DIR = Path("/Users/aniket/Desktop/geetai-template-engine/scripts")

# ---------------------------------------------------------------------------
# 1. Collect all product .ts files (exclude index.ts, registry.ts)
# ---------------------------------------------------------------------------
product_files = sorted(
    p for p in PRODUCTS_DIR.glob("*.ts")
    if p.name not in ("index.ts", "registry.ts")
)

# ---------------------------------------------------------------------------
# 2. Collect all SVG files on disk
# ---------------------------------------------------------------------------
all_svg_files = sorted(f.name for f in IMAGES_DIR.glob("*.svg"))

# ---------------------------------------------------------------------------
# 3. Regex helpers
# ---------------------------------------------------------------------------
RE_SLUG = re.compile(r'^\s+slug:\s*"([^"]+)"', re.MULTILINE)
RE_PRODUCT_NAME = re.compile(r'^\s+product:\s*"([^"]+)"', re.MULTILINE)
RE_BRAND = re.compile(r'^\s+brand:\s*"([^"]+)"', re.MULTILINE)
RE_PRICE = re.compile(r'^\s+price:\s*(\d+)', re.MULTILINE)
RE_ORIGINAL_PRICE = re.compile(r'^\s+originalPrice:\s*(\d+)', re.MULTILINE)
RE_RATING = re.compile(r'^\s+rating:\s*([\d.]+)', re.MULTILINE)
RE_REVIEW_COUNT = re.compile(r'^\s+reviewCount:\s*(\d+)', re.MULTILINE)
RE_CATEGORY = re.compile(r'^\s+category:\s*"([^"]+)"', re.MULTILINE)
RE_IMAGE_SRC = re.compile(r'\bsrc:\s*"([^"]+)"')
RE_BUYLINK_URL = re.compile(r'\burl:\s*"([^"]+)"', re.MULTILINE)

# Known slug/filename mismatches
KNOWN_SLUG_FILENAME_ISSUES = {
    "macbook-pro.ts": ("macbook-pro-16-m4", "macbook-pro"),
    "logitech-mx-master-4.ts": ("logitech-mx-master-4s", "logitech-mx-master-4"),
    "samsung-qd-oled.ts": ("samsung-qd-oled-49", "samsung-qd-oled"),
}

# Generic store domains (no affiliate params)
GENERIC_DOMAINS = [
    "amazon.com", "bestbuy.com", "bhphoto.com", "adorama.com",
    "walmart.com", "gamestop.com", "verizon.com", "samsung.com",
    "apple.com", "sony.com", "nintendo.com", "logitech.com",
    "dji.com", "meta.com", "google.com", "store.google.com",
]

def extract_string(content: str, regex) -> str | None:
    m = regex.search(content)
    return m.group(1) if m else None

def extract_int(content: str, regex) -> int | None:
    m = regex.search(content)
    return int(m.group(1)) if m else None

def extract_float(content: str, regex) -> float | None:
    m = regex.search(content)
    return float(m.group(1)) if m else None

def extract_all(content: str, regex) -> list[str]:
    return regex.findall(content)

def filename_to_expected_slug(filename: str) -> str:
    """Convert 'macbook-pro.ts' → 'macbook-pro'"""
    return filename.replace(".ts", "")

def check_affiliate_url(url: str) -> bool:
    """A URL is affiliate if it contains known affiliate params"""
    affiliate_params = ["ref=", "tag=", "partner=", "affiliate=", "aff=", "referral="]
    return any(p in url.lower() for p in affiliate_params)

def strip_image_prefix(src: str) -> str:
    """Convert '/images/xyz.svg' → 'xyz.svg'"""
    return src.replace("/images/", "").lstrip("/")

# ---------------------------------------------------------------------------
# 4. Parse every product file
# ---------------------------------------------------------------------------
rows = []
all_referenced_images = set()
summary_issues = []

for pf in product_files:
    filename = pf.name
    content = pf.read_text(encoding="utf-8")

    slug = extract_string(content, RE_SLUG)
    product_name = extract_string(content, RE_PRODUCT_NAME)
    brand = extract_string(content, RE_BRAND)
    price = extract_int(content, RE_PRICE)
    orig_price = extract_int(content, RE_ORIGINAL_PRICE)
    rating = extract_float(content, RE_RATING)
    review_count = extract_int(content, RE_REVIEW_COUNT)
    category = extract_string(content, RE_CATEGORY)

    # Image src paths
    image_srcs = extract_all(content, RE_IMAGE_SRC)  # e.g. "/images/xyz.svg"

    # BuyLink urls
    buy_urls = extract_all(content, RE_BUYLINK_URL)

    # --- Slug vs Filename ---
    expected_slug = filename_to_expected_slug(filename)
    slug_matches = slug == expected_slug

    if not slug_matches:
        summary_issues.append(f"SLUG: {filename} → slug '{slug}' (expected '{expected_slug}')")

    # --- Images cross-reference ---
    referenced_files = set()
    missing_images = []
    found_images = []

    for src in image_srcs:
        rel = strip_image_prefix(src)
        referenced_files.add(rel)
        if rel in all_svg_files:
            found_images.append(rel)
        else:
            missing_images.append(rel)

    all_referenced_images |= referenced_files

    images_found = len(found_images)
    images_ref_count = len(referenced_files)

    # --- Extra images (on disk but not referenced by this product) ---
    # "Extra" for a product means any svg file that COULD belong to it but is not referenced.
    # We'll compute global extra images later across all products.

    # --- BuyLinks analysis ---
    total_buylinks = len(buy_urls)
    affiliate_count = sum(1 for u in buy_urls if check_affiliate_url(u))
    affiliate_ready = "Y" if affiliate_count >= (total_buylinks // 2 + 1) else "N"

    # --- Notes ---
    notes = []
    if not orig_price:
        notes.append("Missing originalPrice")
    if not slug_matches:
        notes.append(f"Slug '{slug}' != filename '{expected_slug}'")
    if missing_images:
        notes.append(f"Missing images: {', '.join(missing_images)}")

    rows.append({
        "filename": filename,
        "slug": slug or "",
        "product_name": product_name or "",
        "brand": brand or "",
        "category": category or "",
        "price": price or 0,
        "originalPrice": orig_price or 0,
        "has_originalPrice": "YES" if orig_price is not None else "NO",
        "rating": rating or 0.0,
        "reviewCount": review_count or 0,
        "slug_matches_filename": "YES" if slug_matches else "NO",
        "images_referenced_count": images_ref_count,
        "images_found_count": images_found,
        "images_missing": ", ".join(missing_images) if missing_images else "",
        "images_extra": "",  # filled later
        "buyLinks_total": total_buylinks,
        "buyLinks_affiliate_count": affiliate_count,
        "buyLinks_affiliate_ready": affiliate_ready,
        "buyLinks_urls": ", ".join(buy_urls),
        "notes": "; ".join(notes),
    })

# ---------------------------------------------------------------------------
# 5. Compute global extra images
# ---------------------------------------------------------------------------
# Extra images = SVG files on disk that NO product references
all_expected_image_bases = {}  # mapping from product filename to the expected base name
for pf in product_files:
    base = pf.name.replace(".ts", "")
    all_expected_image_bases[pf.name] = base

referenced_across_all = set()
for row in rows:
    for src in re.findall(r'([\w-]+\.svg)', row["buyLinks_urls"] + row["images_missing"] + row.get("images_found", "")):
        pass  # Not needed, we use all_referenced_images

# Actually, let's be more precise: track every referenced SVG filename
all_ref_svg = set()
for row in rows:
    for img in row["images_missing"].split(", "):
        if img:
            all_ref_svg.add(img)
    # Also include found ones by re-parsing from the file data
for pf in product_files:
    content = pf.read_text(encoding="utf-8")
    for src in extract_all(content, RE_IMAGE_SRC):
        all_ref_svg.add(strip_image_prefix(src))

# We also need to re-map missing images that exist on disk but with different names
# e.g. 'mx-master-4-angle.svg' vs 'logitech-mx-master-4-angle.svg'
# The script already reports these as missing, but for "extra" computation,
# we should consider the DISK files that have zero references across all products.
# If logitech-mx-master-4.ts references 'mx-master-4-angle.svg' (doesn't exist as that name)
# but 'logitech-mx-master-4-angle.svg' exists on disk, the disk file IS referenced by a product
# (just not by the exact filename the product declared), so it's not truly "extra."

# So "extra" = SVG on disk NOT matched by ANY product's image src (after accounting for the
# real filename on disk). This is tricky. Let's do a simpler approach:
# For each SVG on disk, check if ANY product's images array references it (by exact name match).
# If not, it's extra.

extra_svgs = []
for svg in all_svg_files:
    if svg not in all_ref_svg:
        extra_svgs.append(svg)

# Update rows with extra images (same for all rows since it's a global count)
for row in rows:
    row["images_extra"] = ", ".join(extra_svgs)

# Also add an "extra_images" note
for row in rows:
    if extra_svgs and "Extra" not in row["notes"]:
        row["notes"] += f"; Extra images: {', '.join(extra_svgs)}" if row["notes"] else f"Extra images: {', '.join(extra_svgs)}"

# ---------------------------------------------------------------------------
# 6. Generate XLSX
# ---------------------------------------------------------------------------
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
xlsx_path = OUTPUT_DIR / "CURRENT_PRODUCTS_AUDIT.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Products Audit"

# Column headers
headers = [
    "filename", "slug", "product_name", "brand", "category",
    "price", "originalPrice", "has_originalPrice", "rating", "reviewCount",
    "slug_matches_filename", "images_referenced_count", "images_found_count",
    "images_missing", "images_extra",
    "buyLinks_total", "buyLinks_affiliate_count", "buyLinks_affiliate_ready",
    "buyLinks_urls", "notes",
]

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for row_idx, row_data in enumerate(rows, 2):
    for col_idx, header in enumerate(headers, 1):
        val = row_data[header]
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Highlight problematic cells
        if header == "slug_matches_filename" and val == "NO":
            cell.fill = red_fill
        elif header == "slug_matches_filename" and val == "YES":
            cell.fill = green_fill
        if header == "has_originalPrice" and val == "NO":
            cell.fill = red_fill
        if header == "buyLinks_affiliate_ready" and val == "N":
            cell.fill = red_fill

# Auto-width (approximate)
col_widths = {
    1: 30, 2: 30, 3: 30, 4: 15, 5: 15,
    6: 10, 7: 10, 8: 16, 9: 8, 10: 12,
    11: 20, 12: 22, 13: 18,
    14: 40, 15: 40,
    16: 14, 17: 22, 18: 22,
    19: 60, 20: 60,
}
for col_idx, width in col_widths.items():
    ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = width

# Freeze top row
ws.freeze_panes = "A2"

wb.save(xlsx_path)
print(f"✅ Audit saved to: {xlsx_path}")

# ---------------------------------------------------------------------------
# 7. Console summary
# ---------------------------------------------------------------------------
total_products = len(rows)
total_no_orig_price = sum(1 for r in rows if r["has_originalPrice"] == "NO")
total_slug_mismatch = sum(1 for r in rows if r["slug_matches_filename"] == "NO")
total_missing_images = sum(1 for r in rows if r["images_missing"])
total_not_affiliate = sum(1 for r in rows if r["buyLinks_affiliate_ready"] == "N")

print(f"\n{'='*70}")
print(f"  PHASE 1 AUDIT SUMMARY")
print(f"{'='*70}")
print(f"  Products audited:      {total_products}")
print(f"  Total images on disk:  {len(all_svg_files)}")
print(f"  Extra images (unref):  {len(extra_svgs)}")
print(f"{'='*70}")
print(f"  ISSUES FOUND:")
print(f"  - Missing originalPrice:    {total_no_orig_price}")
print(f"  - Slug/filename mismatch:   {total_slug_mismatch}")
print(f"  - Missing images:           {total_missing_images}")
print(f"  - No affiliate links:       {total_not_affiliate}")
print(f"{'='*70}\n")

if total_no_orig_price:
    print("  ❌ Products missing originalPrice:")
    for r in rows:
        if r["has_originalPrice"] == "NO":
            print(f"     - {r['filename']} ({r['slug']})")

if total_slug_mismatch:
    print("\n  ❌ Slug/filename mismatches:")
    for r in rows:
        if r["slug_matches_filename"] == "NO":
            print(f"     - {r['filename']} → slug '{r['slug']}'")

if total_missing_images:
    print("\n  ❌ Products with missing images:")
    for r in rows:
        if r["images_missing"]:
            print(f"     - {r['filename']}: {r['images_missing']}")

if extra_svgs:
    print(f"\n  📦 Extra images (on disk, never referenced): {len(extra_svgs)}")
    for svg in extra_svgs:
        print(f"     - {svg}")

if total_not_affiliate:
    print(f"\n  ⚠️  Products with mostly generic (non-affiliate) buyLinks:")
    for r in rows:
        if r["buyLinks_affiliate_ready"] == "N":
            print(f"     - {r['filename']}: {r['buyLinks_affiliate_count']}/{r['buyLinks_total']} affiliate URLs")

print(f"\n{'='*70}")
print(f"  ✅ Audit complete: {xlsx_path}")
print(f"{'='*70}")
