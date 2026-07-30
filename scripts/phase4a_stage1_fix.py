#!/usr/bin/env python3
"""
Phase 4A — Stage 1: SVG Image Reference Audit & Fix
- Reads each product .ts file
- Cross-references image src paths against actual SVGs on disk
- Maps broken refs to correct existing SVG files
- Fixes the .ts files
- Generates IMAGE_AUDIT_REPORT.xlsx
"""

import json
import os
import re
import subprocess
import sys
from pathlib import Path

BASE_DIR = Path("/Users/aniket/Desktop/geetai-template-engine")
PRODUCTS_DIR = BASE_DIR / "src" / "data" / "products"
IMAGES_DIR = BASE_DIR / "public" / "images"
OUTPUT_DIR = BASE_DIR / "output"

# ── 1. Get all actual SVG files on disk ──────────────────────────────────
actual_svgs = set()
for f in IMAGES_DIR.glob("*.svg"):
    actual_svgs.add(f.name)

print(f"Total SVG files on disk: {len(actual_svgs)}")

# ── 2. Product base names per filename ────────────────────────────────────
PRODUCT_BASES = {
    "airpods-pro-3.ts": "airpods-pro-3",
    "apple-studio-display-2.ts": "apple-studio-display-2",
    "apple-watch-ultra-3.ts": "apple-watch-ultra-3",
    "dell-xps-16-2025.ts": "dell-xps-16-2025",
    "dji-air-4.ts": "dji-air-4",
    "galaxy-s25-ultra.ts": "galaxy-s25-ultra",
    "google-pixel-9-pro.ts": "google-pixel-9-pro",
    "ipad-pro-13-m4.ts": "ipad-pro-13-m4",
    "iphone-16-pro-max.ts": "iphone-16-pro-max",
    "kindle-scribe-2.ts": "kindle-scribe-2",
    "logitech-mx-master-4.ts": "logitech-mx-master-4",
    "macbook-air-15-m3.ts": "macbook-air-15-m3",
    "macbook-pro.ts": "macbook-pro",
    "meta-quest-4.ts": "meta-quest-4",
    "nintendo-switch-2.ts": "nintendo-switch-2",
    "ps5-pro.ts": "ps5-pro",
    "samsung-qd-oled.ts": "samsung-qd-oled",
    "sonos-era-300.ts": "sonos-era-300",
    "sony-a7v.ts": "sony-a7v",
    "sony-wh-1000xm6.ts": "sony-wh-1000xm6",
}

# Standard view types available for every product
VIEWS = ["front", "angle", "side", "cover", "display"]
ALT_MAP = {
    "front": {"front", "hero", "main"},
    "angle": {"angle", "angled"},
    "side": {"side", "ear", "gimbal", "controller", "joycon"},
    "cover": {"cover", "case", "packaging", "box"},
    "display": {"display", "screen", "charging", "passthrough", "top", "back", "lens", "keyboard", "magic-keyboard", "magic_keyboard", "vertical", "camera", "band", "controllers", "docked", "writing", "split"},
}

def best_view_for(ref_name, product_base):
    """Map a reference filename to the best actual SVG view."""
    ref_stem = ref_name.replace(".svg", "")
    # Strip the product base prefix if present
    for base in sorted(PRODUCT_BASES.values(), key=len, reverse=True):
        if ref_stem.startswith(base + "-"):
            view_part = ref_stem[len(base) + 1:]
            break
        if ref_stem == base + "-back":
            return "cover"
    else:
        # Try to guess from the reference name itself
        view_part = ref_stem.split("-")[-1]
    
    view_part_lower = view_part.lower()
    
    # Map to a standard view
    for std_view, aliases in ALT_MAP.items():
        if view_part_lower in aliases:
            return std_view
    
    # Fuzzy match: check if any alias is in the filename
    for std_view, aliases in ALT_MAP.items():
        for alias in aliases:
            if alias in ref_stem.lower():
                return std_view
    
    return "front"  # fallback

def deduce_best_image(reference, product_base):
    """From a broken reference, determine the best actual SVG path."""
    # Get the view part
    ref_filename = reference.lstrip("/images/")
    ref_stem = ref_filename.replace(".svg", "")
    
    # Strip product prefix if any to get view
    view_part = None
    for base in sorted(PRODUCT_BASES.values(), key=len, reverse=True):
        if ref_stem.startswith(base + "-"):
            view_part = ref_stem[len(base) + 1:]
            break
    
    if view_part is None:
        # If product slug doesn't match (e.g., "mx-master-4" vs "logitech-mx-master-4")
        # Try to match against the ref stem patterns
        view_part = ref_stem.split("-")[-1] if "-" in ref_stem else "front"
    
    # Map view_part to a standard view
    std_view = best_view_for(ref_filename, product_base)
    
    return f"/images/{product_base}-{std_view}.svg"

# ── 3. Read product files and build fix plan ─────────────────────────────
results = []
total_refs = 0
total_broken = 0
total_fixed = 0

for filename in sorted(PRODUCT_BASES.keys()):
    product_base = PRODUCT_BASES[filename]
    filepath = PRODUCTS_DIR / filename
    
    content = filepath.read_text()
    
    # Extract images array using regex
    images_match = re.search(r'images:\s*\[(.*?)\]', content, re.DOTALL)
    if not images_match:
        results.append({
            "filename": filename,
            "base": product_base,
            "status": "ERROR",
            "refs": [],
            "broken": [],
            "notes": "Could not parse images array"
        })
        continue
    
    images_block = images_match.group(1)
    
    # Extract src values
    srcs = re.findall(r'src:\s*"([^"]+)"', images_block)
    
    product_refs = []
    product_broken = []
    fixes_needed = []
    
    for src in srcs:
        svg_filename = src.lstrip("/images/")
        total_refs += 1
        
        exists = svg_filename in actual_svgs
        
        product_refs.append({
            "src": src,
            "svg": svg_filename,
            "exists": exists
        })
        
        if not exists:
            total_broken += 1
            # Deduce the correct path
            correct = deduce_best_image(src, product_base)
            correct_svg = correct.lstrip("/images/")
            # Verify it actually exists
            if correct_svg in actual_svgs:
                product_broken.append({
                    "old": src,
                    "new": correct
                })
                fixes_needed.append((src, correct))
                total_fixed += 1
            else:
                # Try even harder - iterate all SVGs for this product
                product_svgs = [s for s in actual_svgs if s.startswith(product_base + "-")]
                if product_svgs:
                    # Use first available as fallback
                    fallback = f"/images/{product_svgs[0]}"
                    product_broken.append({
                        "old": src,
                        "new": fallback
                    })
                    fixes_needed.append((src, fallback))
    
    # Store result
    results.append({
        "filename": filename,
        "base": product_base,
        "status": "OK" if not product_broken else "BROKEN",
        "refs": product_refs,
        "broken": product_broken,
        "fixes_needed": fixes_needed,
        "notes": f"{len(product_broken)} broken refs" if product_broken else "All refs OK"
    })

# ── 4. Apply fixes to .ts files ──────────────────────────────────────────
print(f"\nTotal references: {total_refs}")
print(f"Total broken: {total_broken}")
print(f"Total fixable: {total_fixed}")
print()

fixed_count = 0
for r in results:
    if not r["fixes_needed"]:
        continue
    filepath = PRODUCTS_DIR / r["filename"]
    content = filepath.read_text()
    
    for old, new in r["fixes_needed"]:
        old_escaped = re.escape(old)
        if re.search(old_escaped, content):
            content = re.sub(old_escaped, new, content)
            fixed_count += 1
            print(f"  FIX: {r['filename']}: {old} -> {new}")
    
    filepath.write_text(content)
    r["fix_applied"] = True

print(f"\n✅ Applied {fixed_count} image path fixes across {sum(1 for r in results if r['fixes_needed'])} files")

# ── 5. Generate IMAGE_AUDIT_REPORT.xlsx ─────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Audit"
    
    # Headers
    headers = [
        "Filename", "Product Base", "Status", "Total Refs",
        "Correct Refs", "Broken Refs", "Fixes Applied",
        "Image Paths (before fix)", "Notes"
    ]
    
    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["filename"])
        ws.cell(row=row_idx, column=2, value=r["base"])
        ws.cell(row=row_idx, column=3, value=r["status"])
        ws.cell(row=row_idx, column=4, value=len(r["refs"]))
        ws.cell(row=row_idx, column=5, value=sum(1 for ref in r["refs"] if ref["exists"]))
        ws.cell(row=row_idx, column=6, value=len(r["broken"]))
        ws.cell(row=row_idx, column=7, value="Yes" if r.get("fix_applied") else "No")
        ws.cell(row=row_idx, column=8, value="; ".join(ref["src"] for ref in r["refs"]))
        ws.cell(row=row_idx, column=9, value=r["notes"])
        
        # Color rows
        row_fill = green_fill if r["status"] == "OK" else red_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = row_fill
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, len(results) + 2):
            val = str(ws.cell(row=row, column=col).value or "")
            if len(val) > max_len:
                max_len = min(len(val), 80)
        ws.column_dimensions[chr(64 + col) if col <= 26 else "?"].width = max_len + 2
    
    # Sheet 2: SVG Inventory
    ws2 = wb.create_sheet("SVG Inventory")
    ws2.cell(row=1, column=1, value="SVG Filename")
    ws2.cell(row=1, column=2, value="Referenced By")
    ws2.cell(row=1, column=3, value="Referenced")
    
    for row_idx, svg in enumerate(sorted(actual_svgs), 2):
        # Find which products reference it
        referencing = [r["filename"] for r in results if any(ref["svg"] == svg for ref in r["refs"])]
        # Also check fixed refs
        for r in results:
            if r.get("fix_applied"):
                for broken in r["broken"]:
                    new_svg = broken["new"].lstrip("/images/")
                    if new_svg == svg:
                        if r["filename"] not in referencing:
                            referencing.append(r["filename"] + " (fixed)")
        
        ws2.cell(row=row_idx, column=1, value=svg)
        ws2.cell(row=row_idx, column=2, value=", ".join(referencing) if referencing else "UNREFERENCED")
        ws2.cell(row=row_idx, column=3, value="Yes" if referencing else "No")
        
        if not referencing:
            for col in range(1, 4):
                ws2.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3.cell(row=1, column=1, value="Metric")
    ws3.cell(row=1, column=2, value="Value")
    
    ok_count = sum(1 for r in results if r["status"] == "OK")
    broken_count = sum(1 for r in results if r["status"] == "BROKEN")
    
    metrics = [
        ("Total Products", len(results)),
        ("Products with OK Images", ok_count),
        ("Products with Broken Images", broken_count),
        ("Total Image References", total_refs),
        ("Broken References Found", total_broken),
        ("Fixes Applied", fixed_count),
        ("Unreferenced SVGs", sum(1 for svg in actual_svgs if not any(ref["svg"] == svg for r in results for ref in r["refs"]))),
    ]
    
    for row_idx, (metric, value) in enumerate(metrics, 2):
        ws3.cell(row=row_idx, column=1, value=metric)
        ws3.cell(row=row_idx, column=2, value=value)
    
    xlsx_path = OUTPUT_DIR / "IMAGE_AUDIT_REPORT.xlsx"
    wb.save(str(xlsx_path))
    print(f"\n📊 IMAGE_AUDIT_REPORT.xlsx saved to {xlsx_path}")
    
except ImportError:
    print("openpyxl not available - skipping XLSX generation")

# ── 6. Summary ──────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("PHASE 4A — STAGE 1 COMPLETE")
print("=" * 60)
print(f"Total products audited: {len(results)}")
print(f"Products with correct SVG refs: {ok_count}")
print(f"Products with broken SVG refs: {broken_count}")
print(f"Total broken refs found: {total_broken}")
print(f"Fixes applied: {fixed_count}")
print(f"Unreferenced SVGs: {sum(1 for svg in actual_svgs if not any(ref['svg'] == svg for r in results for ref in r['refs']))}")
print()

# List fixes
for r in results:
    if r.get("fix_applied"):
        for b in r["broken"]:
            print(f"  ✅ {r['filename']}: {b['old'].split('/')[-1]} → {b['new'].split('/')[-1]}")

print()
print("Remaining issues to fix manually:")
for r in results:
    if r["status"] == "BROKEN" and not r.get("fix_applied"):
        for ref in r["refs"]:
            if not ref["exists"]:
                print(f"  ❌ {r['filename']}: {ref['src']} (NO MATCH FOUND)")
